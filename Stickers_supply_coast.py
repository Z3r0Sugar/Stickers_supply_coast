import requests
import time
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# === Загрузка справочника ===
stickers_reference = pd.read_excel("Stickers.xlsx")
stickers_reference = stickers_reference.rename(columns={
    "By": "Коллекция",
    "Collections": "Сабколлекция"
})
stickers_reference["Коллекция"] = stickers_reference["Коллекция"].astype(str)
stickers_reference["Сабколлекция"] = stickers_reference["Сабколлекция"].astype(str)

# === Настройка ===
os.makedirs("log", exist_ok=True)
error_log = []
result_rows = []

def read_x_user_data():
    with open("user_data.txt", encoding="utf-8") as f:
        return f.read().strip()

def get_all_collections(headers):
    try:
        url = "https://palacenft.com/api/v1/markets/collections?onSale=true"
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"❌ Ошибка при получении коллекций: {e}")
        return []

def get_packs(collection_id, headers, retries=2):
    url = f"https://palacenft.com/api/v1/markets/packs?collection_id={collection_id}"
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=headers, timeout=10)
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            print(f"⚠️ Попытка {attempt+1}: ошибка packs для коллекции {collection_id}: {e}")
            time.sleep(2)
    return []

def get_floor_price(collection_id, pack_id, headers, retries=3):
    url = (
        f"https://palacenft.com/api/v1/markets/offers"
        f"?collection_id={collection_id}&pack_id={pack_id}&limit=40&offset=0&sort=price_asc"
    )
    session = requests.Session()
    session.headers.update(headers)
    session.keep_alive = False

    for attempt in range(retries):
        try:
            resp = session.get(url, timeout=15)
            resp.raise_for_status()
            offers = resp.json().get("offers", [])
            if offers:
                return round(float(offers[0]["price"]), 2)
            return None
        except Exception as e:
            print(f"⚠️ Ошибка floor для pack_id={pack_id}: {e}")
            time.sleep(5)
    return None

# === Главная логика ===

def main():
    x_user_data = read_x_user_data()
    headers = {
        "x-user-data": x_user_data,
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0"
    }

    collections = get_all_collections(headers)
    print(f" Найдено {len(collections)} коллекций\n")

    for col in collections:
        col_id = col["id"]
        col_name = col["name"]
        print(f"\n Коллекция: {col_name} (id={col_id})")

        packs = get_packs(col_id, headers)
        time.sleep(0.2)

        for pack in packs:
            pack_name = pack["name"]
            pack_id = pack["id"]
            print(f"  └─ Сабколлекция: {pack_name} (pack_id={pack_id})")

            price = get_floor_price(col_id, pack_id, headers)
            time.sleep(0.2)

            if price is not None:
                print(f"      Floor: {price:.2f} TON")
                row = {
                    "Коллекция": col_name,
                    "Сабколлекция": pack_name,
                    "Floor (TON)": price  # оставим числом, форматируем позже
                }

                match = stickers_reference[
                    (stickers_reference["Коллекция"].str.strip().str.lower() == col_name.strip().lower()) &
                    (stickers_reference["Сабколлекция"].str.strip().str.lower() == str(pack_name).strip().lower())
                ]

                if not match.empty:
                    match_row = match.iloc[0]
                    for field in ["Initial price (stars)", "Initial price ($)", "Issued", "Date"]:
                        row[field] = match_row.get(field, "")

                result_rows.append(row)
            else:
                print(f"     ⚠️ Нет офферов")

    # === Excel-сохранение ===
    if result_rows:
        df = pd.DataFrame(result_rows)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"floor_prices_{timestamp}.xlsx"
        df.to_excel(filename, index=False)

        # === Форматирование Excel ===
        wb = load_workbook(filename)
        ws = wb.active

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(horizontal="center", vertical="center")

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in col_cells:
                cell.alignment = align
                cell.border = border
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 2

        # Жирный заголовок
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Формат чисел с запятой
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'

        # Формат даты (столбец Date)
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
            for cell in row:
                cell.number_format = 'DD MMM YYYY'

        wb.save(filename)
        print(f"\n Готово! Сохранено в {filename}")

if __name__ == "__main__":
    main()

input("Парсинг завершён. Для закрытия окна нажмите Enter")

#Create by @zer0_sugar